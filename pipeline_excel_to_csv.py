"""
=============================================================
  PIPELINE — Excel → data_daily.csv
  Ajinomoto FI-2 Production Monitoring
=============================================================
  CARA PAKAI:
  1. Taruh file Excel bulanan di folder INPUT_FOLDER (default: "input/")
  2. Nama file harus mengandung nama bulan & tahun, contoh:
       Januari_2026.xlsx  /  februari_2026.xlsx  /  maret2026.xlsx
       ATAU format: rev140126.xlsx (auto-detect bulan dari isi sheet)
  3. Jalankan: python pipeline_excel_to_csv.py
  4. Output: output/data_daily.csv (siap dipakai Streamlit dashboard)
=============================================================
"""

import os
import re
import glob
import openpyxl
import pandas as pd

# ── CONFIG ───────────────────────────────────────────────────
INPUT_FOLDER  = "input"
OUTPUT_FOLDER = "output"
OUTPUT_FILE   = os.path.join(OUTPUT_FOLDER, "data_daily.csv")

# ── MAPPING ──────────────────────────────────────────────────
BULAN_ID = {
    1:"Januari", 2:"Februari", 3:"Maret", 4:"April",
    5:"Mei", 6:"Juni", 7:"Juli", 8:"Agustus",
    9:"September", 10:"Oktober", 11:"November", 12:"Desember"
}
BULAN_STR = {v.lower(): k for k, v in BULAN_ID.items()}
# Tambahan singkatan
BULAN_STR.update({
    "jan":1,"feb":2,"mar":3,"apr":4,"mei":5,"jun":6,
    "jul":7,"agu":8,"ags":8,"sep":9,"okt":10,"nov":11,"des":12
})

TARGET_MAP = {
    "A5RB": 2016, "Amamiplus": 44, "PTPA 20gr": 1952, "MJCL": 2187,
    "PTCL": 3904, "50G": 1418, "100G": 1575, "Ajiplus": 1260,
    "Ajiplus Ekicho": 1260, "Ajinomoto Plus": 1260, "AJP 200gr": 2016,
    "Ajiplus Premium": 1008, "ekicho 20Kg": 1008, "250RC": 1008,
}

# ── KOLOM SMALL/MEDIUM (col index 1-based) ───────────────────
# (nama_produk, col_daily, col_acc, line, mesin, kg_ctn)
SM_PRODUCTS = [
    ("A5RB",            3,  4,  "LINE-3",  "Vertical Packer",  10.8),
    ("Amamiplus",       5,  6,  "LINE-13", "YDE",              12.0),
    ("PTPA 20gr",       7,  8,  "LINE-11", "UNILOGO",           5.76),
    ("MJCL",            9,  10, "LINE-4",  "GP C",              2.52),
    ("PTCL",            11, 12, "LINE-6",  "GP TT10 E,G,H",    5.76),
    ("50G",             13, 14, "LINE-5",  "GP I & K",         12.0),
    ("100G",            15, 16, "LINE-10", "GP I & K",         12.0),
    ("Ajiplus",         17, 18, "LINE-12", "SANCO",            20.0),
    ("Ajiplus Ekicho",  19, 20, "LINE-12", "SANCO",            20.0),
    ("Ajinomoto Plus",  21, 22, "LINE-12", "SANCO",            20.0),
    ("AJP 200gr",       23, 24, "LINE-12", "SANCO",             6.0),
    ("Ajiplus Premium", 25, 26, "LINE-12", "SANCO",            20.0),
    ("ekicho 20Kg",     27, 28, "LINE-12", "YDE",              20.0),
]

# ── KOLOM BAG TYPE (col index 1-based) ───────────────────────
BAG_PRODUCTS = [
    ("250RC", 31, 32, "LINE-1", "Vertical Packer", 20.0),
    ("250RC", 33, 34, "LINE-2", "Vertical Packer", 20.0),
    ("PTCL",  35, 36, "LINE-8", "GP TT10 E,G,H",  5.76),
    ("50G",   37, 38, "LINE-9", "GP I & K",       12.0),
    ("100G",  39, 40, "LINE-9", "GP I & K",       12.0),
]


def detect_bulan_tahun_from_filename(filename):
    """Deteksi bulan & tahun dari nama file Excel."""
    fname = os.path.basename(filename).lower()
    fname_clean = re.sub(r'[_\-\s\.]+', ' ', fname)

    # Cari tahun (4 digit)
    tahun_match = re.search(r'20\d{2}', fname_clean)
    tahun = int(tahun_match.group()) if tahun_match else None

    # Cari bulan dari nama
    bulan_num = None
    for nama, num in sorted(BULAN_STR.items(), key=lambda x: -len(x[0])):
        if nama in fname_clean:
            bulan_num = num
            break

    # Cari bulan dari format 2 digit angka (misal: 01 = Januari)
    if bulan_num is None:
        m = re.search(r'\b(0?[1-9]|1[0-2])\b', fname_clean)
        if m:
            bulan_num = int(m.group())

    return bulan_num, tahun


def parse_excel_file(filepath, bulan_num, tahun):
    """Parse satu file Excel → DataFrame."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ⚠️  Gagal baca {filepath}: {e}")
        return pd.DataFrame()

    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(min_row=10, max_row=45, values_only=True))

    records = []
    for row in all_rows:
        if len(row) < 30:
            row = list(row) + [None] * (50 - len(row))

        # Ambil tanggal kolom 2 (index 1)
        tgl = row[1]
        if not isinstance(tgl, (int, float)):
            continue
        tgl = int(tgl)
        if not (1 <= tgl <= 31):
            continue

        # ── Small/Medium section ──────────────────────────────
        for produk, c_daily, c_acc, line, mesin, kg_ctn in SM_PRODUCTS:
            if c_acc - 1 >= len(row):
                continue
            daily_val = row[c_daily - 1]
            acc_val   = row[c_acc - 1]
            if daily_val is None and acc_val is None:
                continue
            daily = float(daily_val) if isinstance(daily_val, (int, float)) else 0.0
            acc   = float(acc_val)   if isinstance(acc_val,   (int, float)) else None
            records.append({
                "tahun":      tahun,
                "bulan_num":  bulan_num,
                "bulan":      BULAN_ID[bulan_num],
                "periode":    f"{BULAN_ID[bulan_num]} {tahun}",
                "tgl":        tgl,
                "kategori":   "Small/Medium",
                "produk":     produk,
                "line":       line,
                "mesin":      mesin,
                "kg_ctn":     kg_ctn,
                "target_day": TARGET_MAP.get(produk, 0),
                "daily":      daily,
                "acc":        acc,
            })

        # ── Bag Type section ─────────────────────────────────
        if len(row) < 40:
            continue
        tgl_bag = row[29]  # col 30 = second Tgl column
        if not isinstance(tgl_bag, (int, float)):
            continue

        for produk, c_daily, c_acc, line, mesin, kg_ctn in BAG_PRODUCTS:
            if c_acc - 1 >= len(row):
                continue
            daily_val = row[c_daily - 1]
            acc_val   = row[c_acc - 1]
            if daily_val is None and acc_val is None:
                continue
            daily = float(daily_val) if isinstance(daily_val, (int, float)) else 0.0
            acc   = float(acc_val)   if isinstance(acc_val,   (int, float)) else None
            records.append({
                "tahun":      tahun,
                "bulan_num":  bulan_num,
                "bulan":      BULAN_ID[bulan_num],
                "periode":    f"{BULAN_ID[bulan_num]} {tahun}",
                "tgl":        tgl,
                "kategori":   "Bag Type",
                "produk":     produk,
                "line":       line,
                "mesin":      mesin,
                "kg_ctn":     kg_ctn,
                "target_day": TARGET_MAP.get(produk, 0),
                "daily":      daily,
                "acc":        acc,
            })

    return pd.DataFrame(records)


def run_pipeline():
    print("=" * 60)
    print("  AJINOMOTO FI-2 — Pipeline Excel → data_daily.csv")
    print("=" * 60)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(INPUT_FOLDER,  exist_ok=True)

    # Cari semua file Excel di folder input
    excel_files = (
        glob.glob(os.path.join(INPUT_FOLDER, "*.xlsx")) +
        glob.glob(os.path.join(INPUT_FOLDER, "*.xls"))
    )

    if not excel_files:
        print(f"\n❌ Tidak ada file Excel di folder '{INPUT_FOLDER}/'")
        print(f"   Taruh file Januari_2026.xlsx, Februari_2026.xlsx, dst. di sana.")
        return

    print(f"\n📂 Ditemukan {len(excel_files)} file Excel:")
    all_dfs = []

    for fpath in sorted(excel_files):
        fname = os.path.basename(fpath)
        bulan_num, tahun = detect_bulan_tahun_from_filename(fname)

        if bulan_num is None or tahun is None:
            print(f"  ⚠️  SKIP: {fname} — tidak bisa deteksi bulan/tahun dari nama file")
            print(f"       Rename file ke format: Januari_2026.xlsx")
            continue

        print(f"  ✅ {fname} → {BULAN_ID[bulan_num]} {tahun}")
        df = parse_excel_file(fpath, bulan_num, tahun)

        if df.empty:
            print(f"     ⚠️  Tidak ada data yang berhasil dibaca dari file ini")
            continue

        print(f"     → {len(df)} baris | Produk: {sorted(df['produk'].unique())}")
        all_dfs.append(df)

    if not all_dfs:
        print("\n❌ Tidak ada data yang berhasil diproses.")
        return

    # Gabungkan & sort
    df_all = pd.concat(all_dfs, ignore_index=True)
    df_all = df_all.sort_values(["tahun", "bulan_num", "tgl", "produk"]).reset_index(drop=True)

    # Hapus duplikat (periode + tgl + produk + line)
    before = len(df_all)
    df_all = df_all.drop_duplicates(subset=["tahun","bulan_num","tgl","produk","line"]).reset_index(drop=True)
    after  = len(df_all)
    if before != after:
        print(f"\n  🔄 Deduplikasi: {before - after} baris duplikat dihapus")

    # Simpan
    df_all.to_csv(OUTPUT_FILE, index=False)

    print(f"\n{'='*60}")
    print(f"✅ BERHASIL! Output: {OUTPUT_FILE}")
    print(f"   Total baris : {len(df_all):,}")
    print(f"   Periode     : {sorted(df_all['periode'].unique())}")
    print(f"   Produk      : {sorted(df_all['produk'].unique())}")
    print(f"   Lines       : {sorted(df_all['line'].unique())}")
    print(f"   Kategori    : {sorted(df_all['kategori'].unique())}")
    print(f"\n👉 Sekarang buka Streamlit dashboard dan klik 'Force Regenerate Pivot'")
    print("=" * 60)


if __name__ == "__main__":
    run_pipeline()
